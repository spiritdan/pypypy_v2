import openpyxl
wb=openpyxl.Workbook()
sheet=wb.active
sheet.title='漫威宇宙'
sheet['A1']='漫威宇宙'
rows=[['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
for i in rows:
    sheet.append(i)
wb.save('Marvel.xlsx')

wb = openpyxl.load_workbook('Marvel.xlsx')
sheet=wb['漫威宇宙']
sheetname = wb.sheetnames
print(sheetname)
A1_cell=sheet['A1']
A1_value=A1_cell.value
print(A1_value)